import threading
import time
import win32api
import pyautogui
import math
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QFileDialog, QSizePolicy
from matplotlib import pyplot as plt
import calculateParameters
import pandas as pd
import openpyxl
import mainWindow
from PyQt5.QtWidgets import QTableWidgetItem
import generateMap
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import numpy as np


initial_point = []
degree = 0
distance = 0
pos = 0
last_distance = 0
last_point = []

table_data_launch = [[]]
table_data_impact = [[]]
dist_sensor_launch = []
dist_sensor_impact = []
dist_event = []
bearing_sensor_launch = []
bearing_sensor_impact = []
bearing_event = []

table_pointer = 0
delays = []
delays_total = []
delay_ranking = []
total_differences = []
total_ordered_diff = []
delays_compared = []

time = []
X_pos = []
Y_pos = []
Z_pos = []
Vx_list = []
Vy_list = []
Vz_list = []
time_2 = []
X_pos_2 = []
Y_pos_2 = []
Z_pos_2 = []
Vx_list_2 = []
Vy_list_2 = []
Vz_list_2 = []
t_end = 0

class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(1700, 960)
        self.frame_3 = QtWidgets.QFrame(Dialog)
        self.frame_3.setGeometry(QtCore.QRect(1400, 30, 281, 891))
        self.frame_3.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_3.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.frame_3.setObjectName("frame_3")
        self.frame_4 = QtWidgets.QFrame(Dialog)
        self.frame_4.setGeometry(QtCore.QRect(1030, 30, 361, 891))
        self.frame_4.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_4.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_4.setObjectName("frame_4")
        self.pushButton = QtWidgets.QPushButton(self.frame_4, clicked = lambda: self.exportExcel())
        self.pushButton.setGeometry(QtCore.QRect(10, 830, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(7)
        self.pushButton.setFont(font)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("save.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton.setIcon(icon)
        self.pushButton.setObjectName("pushButton")
        self.pushButton_2 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_2.setGeometry(QtCore.QRect(120, 830, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(7)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setAcceptDrops(False)
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("edit.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_2.setIcon(icon1)
        self.pushButton_2.setFlat(False)
        self.pushButton_2.setObjectName("pushButton_2")
        self.table = QtWidgets.QTableWidget(Dialog)
        self.table.setGeometry(QtCore.QRect(20, 31, 996, 451))
        self.table_2 = QtWidgets.QTableWidget(Dialog)
        self.table_2.setGeometry(QtCore.QRect(20, 31, 996, 451))
        self.table_3 = QtWidgets.QTableWidget(Dialog)
        self.table_3.setGeometry(QtCore.QRect(20, 31, 996, 451))
        self.table_4 = QtWidgets.QTableWidget(Dialog)
        self.table_4.setGeometry(QtCore.QRect(20, 31, 996, 451))
        self.table_5 = QtWidgets.QTableWidget(Dialog)
        self.table_5.setGeometry(QtCore.QRect(20, 31, 996, 451))
        self.table_6 = QtWidgets.QTableWidget(Dialog)
        self.table_6.setGeometry(QtCore.QRect(20, 31, 996, 451))
        self.table_7 = QtWidgets.QTableWidget(Dialog)
        self.table_7.setGeometry(QtCore.QRect(20, 31, 996, 451))
        self.table_8 = QtWidgets.QTableWidget(Dialog)
        self.table_8.setGeometry(QtCore.QRect(20, 31, 996, 451))
        self.table_9 = QtWidgets.QTableWidget(Dialog)
        self.table_9.setGeometry(QtCore.QRect(20, 31, 996, 451))
        self.table_10 = QtWidgets.QTableWidget(Dialog)
        self.table_10.setGeometry(QtCore.QRect(20, 31, 996, 451))
        self.comboBox = QtWidgets.QComboBox(self.frame_4)
        self.comboBox.setGeometry(QtCore.QRect(21, 10, 141, 21))
        self.comboBox.setObjectName("comboBox")
        self.dial = QtWidgets.QDial(self.frame_4)
        self.dial.setGeometry(QtCore.QRect(80, 154, 181, 121))
        self.dial.setMaximum(360)
        self.dial.setSingleStep(5)
        self.dial.setPageStep(18)
        self.dial.setSliderPosition(180)
        self.dial.setOrientation(QtCore.Qt.Horizontal)
        self.dial.setInvertedAppearance(True)
        self.dial.setInvertedControls(True)
        self.dial.setWrapping(True)
        self.dial.setNotchTarget(1.0)
        self.dial.setNotchesVisible(False)
        self.dial.setObjectName("dial")
        self.horizontalSlider = QtWidgets.QSlider(self.frame_4)
        self.horizontalSlider.setGeometry(QtCore.QRect(750, 790, 271, 20))
        self.horizontalSlider.setOrientation(QtCore.Qt.Horizontal)
        self.horizontalSlider.setObjectName("horizontalSlider")
        self.lineEdit = QtWidgets.QLineEdit(self.frame_4)
        self.lineEdit.setGeometry(QtCore.QRect(120, 366, 101, 20))
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.frame_4)
        self.lineEdit_2.setGeometry(QtCore.QRect(180, 394, 91, 20))
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.lineEdit_3 = QtWidgets.QLineEdit(self.frame_4)
        self.lineEdit_3.setGeometry(QtCore.QRect(70, 394, 91, 20))
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.lineEdit_4 = QtWidgets.QLineEdit(self.frame_4)
        self.lineEdit_4.setGeometry(QtCore.QRect(120, 424, 101, 20))
        self.lineEdit_4.setObjectName("lineEdit_4")
        self.pushButton_3 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_3.setGeometry(QtCore.QRect(40, 384, 31, 41))
        self.pushButton_3.setText("")
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap("linksarrow.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_3.setIcon(icon2)
        self.pushButton_3.setIconSize(QtCore.QSize(36, 36))
        self.pushButton_3.setDefault(False)
        self.pushButton_3.setFlat(True)
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_4 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_4.setGeometry(QtCore.QRect(150, 334, 41, 31))
        self.pushButton_4.setText("")
        icon3 = QtGui.QIcon()
        icon3.addPixmap(QtGui.QPixmap("uparrow.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_4.setIcon(icon3)
        self.pushButton_4.setIconSize(QtCore.QSize(36, 36))
        self.pushButton_4.setDefault(False)
        self.pushButton_4.setFlat(True)
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_5 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_5.setGeometry(QtCore.QRect(268, 384, 31, 41))
        self.pushButton_5.setText("")
        icon4 = QtGui.QIcon()
        icon4.addPixmap(QtGui.QPixmap("rightarrow.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_5.setIcon(icon4)
        self.pushButton_5.setIconSize(QtCore.QSize(36, 36))
        self.pushButton_5.setDefault(False)
        self.pushButton_5.setFlat(True)
        self.pushButton_5.setObjectName("pushButton_5")
        self.pushButton_6 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_6.setGeometry(QtCore.QRect(150, 446, 41, 31))
        self.pushButton_6.setText("")
        icon5 = QtGui.QIcon()
        icon5.addPixmap(QtGui.QPixmap("downarrow.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_6.setIcon(icon5)
        self.pushButton_6.setIconSize(QtCore.QSize(36, 36))
        self.pushButton_6.setDefault(False)
        self.pushButton_6.setFlat(True)
        self.pushButton_6.setObjectName("pushButton_6")
        self.label = QtWidgets.QLabel(self.frame_4)
        self.label.setGeometry(QtCore.QRect(71, 100, 221, 231))
        self.label.setText("")
        self.label.setPixmap(QtGui.QPixmap("compass.png"))
        self.label.setObjectName("label")
        self.label_4 = QtWidgets.QLabel(self.frame_4)
        self.label_4.setGeometry(QtCore.QRect(41, 50, 261, 16))
        self.label_4.setAlignment(QtCore.Qt.AlignCenter)
        self.label_4.setObjectName("label_4")
        self.comboBox_2 = QtWidgets.QComboBox(self.frame_4)
        self.comboBox_2.setGeometry(QtCore.QRect(191, 10, 141, 21))
        self.comboBox_2.setObjectName("comboBox_2")
        self.pushButton_7 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_7.setGeometry(QtCore.QRect(230, 830, 111, 41))
        font = QtGui.QFont()
        font.setPointSize(7)
        self.pushButton_7.setFont(font)
        self.pushButton_7.setAcceptDrops(False)
        icon6 = QtGui.QIcon()
        icon6.addPixmap(QtGui.QPixmap("refresh.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_7.setIcon(icon6)
        self.pushButton_7.setFlat(False)
        self.pushButton_7.setObjectName("pushButton_7")
        self.label_6 = QtWidgets.QLabel(self.frame_4)
        self.label_6.setGeometry(QtCore.QRect(150, 204, 47, 13))
        self.label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.label_6.setObjectName("label_6")
        self.label_06 = QtWidgets.QLabel(self.frame_4)
        self.label_06.setGeometry(QtCore.QRect(150, 214, 47, 13))
        self.label_06.setAlignment(QtCore.Qt.AlignCenter)
        self.label_06.setObjectName("label_06")
        self.horizontalSlider_2 = QtWidgets.QSlider(self.frame_4)
        self.horizontalSlider_2.setGeometry(QtCore.QRect(20, 790, 271, 20))
        self.horizontalSlider_2.setOrientation(QtCore.Qt.Horizontal)
        self.horizontalSlider_2.setObjectName("horizontalSlider_2")
        self.label_7 = QtWidgets.QLabel(self.frame_4)
        self.label_7.setGeometry(QtCore.QRect(40, 500, 261, 20))
        self.label_7.setAlignment(QtCore.Qt.AlignCenter)
        self.label_7.setObjectName("label_7")
        self.label_8 = QtWidgets.QLabel(self.frame_4)
        self.label_8.setGeometry(QtCore.QRect(300, 780, 47, 31))
        self.label_8.setAlignment(QtCore.Qt.AlignCenter)
        self.label_8.setObjectName("label_8")
        self.frame_2 = QtWidgets.QFrame(Dialog)
        self.frame_2.setGeometry(QtCore.QRect(525, 490, 491, 431))
        self.frame_2.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_2.setObjectName("frame_2")
        self.horizontalLayoutWidget_2 = QtWidgets.QWidget(self.frame_2)
        self.horizontalLayoutWidget_2.setGeometry(QtCore.QRect(1, 1, 489, 429))
        self.horizontalLayoutWidget_2.setObjectName("horizontalLayoutWidget_2")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget_2)
        self.horizontalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.frame = QtWidgets.QFrame(Dialog)
        self.frame.setGeometry(QtCore.QRect(20, 490, 491, 431))
        self.frame.setFrameShape(QtWidgets.QFrame.Box)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.horizontalLayoutWidget = QtWidgets.QWidget(self.frame)
        self.horizontalLayoutWidget.setGeometry(QtCore.QRect(0, 0, 489, 429))
        self.horizontalLayoutWidget.setObjectName("horizontalLayoutWidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label_5 = QtWidgets.QLabel(self.frame_4)
        self.label_5.setGeometry(QtCore.QRect(71, 560, 221, 231))
        self.label_5.setText("")
        self.label_5.setPixmap(QtGui.QPixmap("compass.png"))
        self.label_5.setObjectName("label_5")
        self.dial_2 = QtWidgets.QDial(self.frame_4)
        self.dial_2.setGeometry(QtCore.QRect(80, 614, 181, 121))
        self.dial_2.setMaximum(360)
        self.dial_2.setSingleStep(5)
        self.dial_2.setPageStep(18)
        self.dial_2.setSliderPosition(0)
        self.dial_2.setOrientation(QtCore.Qt.Horizontal)
        self.dial_2.setInvertedAppearance(True)
        self.dial_2.setInvertedControls(True)
        self.dial_2.setWrapping(True)
        self.dial_2.setNotchTarget(1.0)
        self.dial_2.setNotchesVisible(False)
        self.dial_2.setObjectName("dial_2")
        self.label_9 = QtWidgets.QLabel(self.frame_4)
        self.label_9.setGeometry(QtCore.QRect(150, 630, 47, 16))
        self.label_9.setAlignment(QtCore.Qt.AlignCenter)
        self.label_9.setObjectName("label_9")
        self.label.raise_()
        self.checkBox = QtWidgets.QCheckBox(self.frame_4)
        self.checkBox.setGeometry(QtCore.QRect(140, 70, 70, 17))
        self.checkBox.setObjectName("checkBox")
        self.checkBox_2 = QtWidgets.QCheckBox(self.frame_4)
        self.checkBox_2.setGeometry(QtCore.QRect(140, 520, 70, 17))
        self.checkBox_2.setObjectName("checkBox_2")
        self.lineEdit_5 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_5.setGeometry(QtCore.QRect(165, 60, 100, 20))
        self.lineEdit_5.setObjectName("lineEdit_5")
        self.lineEdit_6 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_6.setGeometry(QtCore.QRect(165, 105, 100, 20))
        self.lineEdit_6.setObjectName("lineEdit_6")
        self.lineEdit_7 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_7.setGeometry(QtCore.QRect(165, 150, 100, 20))
        self.lineEdit_7.setObjectName("lineEdit_7")
        self.lineEdit_8 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_8.setGeometry(QtCore.QRect(165, 195, 100, 20))
        self.lineEdit_8.setObjectName("lineEdit_8")
        self.lineEdit_9 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_9.setGeometry(QtCore.QRect(165, 240, 100, 20))
        self.lineEdit_9.setObjectName("lineEdit_9")
        self.lineEdit_10 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_10.setGeometry(QtCore.QRect(165, 285, 100, 20))
        self.lineEdit_10.setObjectName("lineEdit_10")
        self.lineEdit_11 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_11.setGeometry(QtCore.QRect(165, 330, 100, 20))
        self.lineEdit_11.setObjectName("lineEdit_11")
        self.lineEdit_12 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_12.setGeometry(QtCore.QRect(165, 375, 100, 20))
        self.lineEdit_12.setObjectName("lineEdit_12")
        self.lineEdit_13 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_13.setGeometry(QtCore.QRect(165, 420, 100, 20))
        self.lineEdit_13.setObjectName("lineEdit_13")
        self.lineEdit_14 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_14.setGeometry(QtCore.QRect(165, 465, 100, 20))
        self.lineEdit_14.setObjectName("lineEdit_14")
        self.lineEdit_15 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_15.setGeometry(QtCore.QRect(165, 735, 100, 20))
        self.lineEdit_15.setObjectName("lineEdit_15")
        self.lineEdit_16 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_16.setGeometry(QtCore.QRect(165, 780, 100, 20))
        self.lineEdit_16.setObjectName("lineEdit_16")
        self.lineEdit_17 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_17.setGeometry(QtCore.QRect(165, 600, 100, 20))
        self.lineEdit_17.setObjectName("lineEdit_17")
        self.lineEdit_18 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_18.setGeometry(QtCore.QRect(165, 690, 100, 20))
        self.lineEdit_18.setObjectName("lineEdit_18")
        self.lineEdit_21 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_21.setGeometry(QtCore.QRect(165, 645, 100, 20))
        self.lineEdit_21.setObjectName("lineEdit_21")
        self.lineEdit_23 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_23.setGeometry(QtCore.QRect(165, 510, 100, 20))
        self.lineEdit_23.setObjectName("lineEdit_23")
        self.lineEdit_24 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_24.setGeometry(QtCore.QRect(165, 555, 100, 20))
        self.lineEdit_24.setObjectName("lineEdit_24")
        self.label_11 = QtWidgets.QLabel(self.frame_3)
        self.label_11.setGeometry(QtCore.QRect(20, 60, 121, 21))
        self.label_11.setAlignment(QtCore.Qt.AlignCenter)
        self.label_11.setObjectName("label_11")
        self.label_12 = QtWidgets.QLabel(self.frame_3)
        self.label_12.setGeometry(QtCore.QRect(20, 105, 121, 21))
        self.label_12.setAlignment(QtCore.Qt.AlignCenter)
        self.label_12.setObjectName("label_12")
        self.label_13 = QtWidgets.QLabel(self.frame_3)
        self.label_13.setGeometry(QtCore.QRect(20, 195, 121, 21))
        self.label_13.setAlignment(QtCore.Qt.AlignCenter)
        self.label_13.setObjectName("label_13")
        self.label_14 = QtWidgets.QLabel(self.frame_3)
        self.label_14.setGeometry(QtCore.QRect(20, 150, 135, 21))
        self.label_14.setAlignment(QtCore.Qt.AlignCenter)
        self.label_14.setObjectName("label_14")
        self.label_15 = QtWidgets.QLabel(self.frame_3)
        self.label_15.setGeometry(QtCore.QRect(20, 375, 121, 21))
        self.label_15.setAlignment(QtCore.Qt.AlignCenter)
        self.label_15.setObjectName("label_15")
        self.label_16 = QtWidgets.QLabel(self.frame_3)
        self.label_16.setGeometry(QtCore.QRect(20, 330, 121, 21))
        self.label_16.setAlignment(QtCore.Qt.AlignCenter)
        self.label_16.setObjectName("label_16")
        self.label_17 = QtWidgets.QLabel(self.frame_3)
        self.label_17.setGeometry(QtCore.QRect(20, 285, 131, 21))
        self.label_17.setAlignment(QtCore.Qt.AlignCenter)
        self.label_17.setObjectName("label_17")
        self.label_18 = QtWidgets.QLabel(self.frame_3)
        self.label_18.setGeometry(QtCore.QRect(20, 420, 121, 21))
        self.label_18.setAlignment(QtCore.Qt.AlignCenter)
        self.label_18.setObjectName("label_18")
        self.label_19 = QtWidgets.QLabel(self.frame_3)
        self.label_19.setGeometry(QtCore.QRect(20, 555, 121, 21))
        self.label_19.setAlignment(QtCore.Qt.AlignCenter)
        self.label_19.setObjectName("label_19")
        self.label_20 = QtWidgets.QLabel(self.frame_3)
        self.label_20.setGeometry(QtCore.QRect(20, 730, 121, 41))
        self.label_20.setAlignment(QtCore.Qt.AlignCenter)
        self.label_20.setWordWrap(True)
        self.label_20.setObjectName("label_20")
        self.label_21 = QtWidgets.QLabel(self.frame_3)
        self.label_21.setGeometry(QtCore.QRect(20, 680, 121, 41))
        self.label_21.setAlignment(QtCore.Qt.AlignCenter)
        self.label_21.setWordWrap(True)
        self.label_21.setObjectName("label_21")
        self.label_22 = QtWidgets.QLabel(self.frame_3)
        self.label_22.setGeometry(QtCore.QRect(10, 780, 151, 21))
        self.label_22.setAlignment(QtCore.Qt.AlignCenter)
        self.label_22.setObjectName("label_22")
        self.label_23 = QtWidgets.QLabel(self.frame_3)
        self.label_23.setGeometry(QtCore.QRect(20, 510, 121, 21))
        self.label_23.setAlignment(QtCore.Qt.AlignCenter)
        self.label_23.setObjectName("label_23")
        self.label_24 = QtWidgets.QLabel(self.frame_3)
        self.label_24.setGeometry(QtCore.QRect(20, 465, 121, 21))
        self.label_24.setAlignment(QtCore.Qt.AlignCenter)
        self.label_24.setObjectName("label_24")
        self.label_25 = QtWidgets.QLabel(self.frame_3)
        self.label_25.setGeometry(QtCore.QRect(20, 630, 121, 41))
        self.label_25.setAlignment(QtCore.Qt.AlignCenter)
        self.label_25.setWordWrap(True)
        self.label_25.setObjectName("label_25")
        self.label_26 = QtWidgets.QLabel(self.frame_3)
        self.label_26.setGeometry(QtCore.QRect(20, 600, 121, 21))
        self.label_26.setAlignment(QtCore.Qt.AlignCenter)
        self.label_26.setObjectName("label_26")
        self.label_27 = QtWidgets.QLabel(self.frame_3)
        self.label_27.setGeometry(QtCore.QRect(20, 240, 121, 21))
        self.label_27.setAlignment(QtCore.Qt.AlignCenter)
        self.label_27.setObjectName("label_27")
        self.label_28 = QtWidgets.QLabel(self.frame_3)
        self.label_28.setGeometry(QtCore.QRect(80, 10, 131, 31))
        self.pushButton_8 = QtWidgets.QPushButton(self.frame_3)
        self.pushButton_8.setGeometry(QtCore.QRect(35, 830, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(7)
        self.pushButton_8.setFont(font)
        icon7 = QtGui.QIcon()
        icon7.addPixmap(QtGui.QPixmap("savesettings.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_8.setIcon(icon7)
        self.pushButton_8.setIconSize(QtCore.QSize(20, 20))
        self.pushButton_8.setObjectName("pushButton_8")
        self.pushButton_9 = QtWidgets.QPushButton(self.frame_3)
        self.pushButton_9.setGeometry(QtCore.QRect(145, 830, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(7)
        self.pushButton_9.setFont(font)
        icon8 = QtGui.QIcon()
        icon8.addPixmap(QtGui.QPixmap("defaultsettings.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_9.setIcon(icon8)
        self.pushButton_9.setIconSize(QtCore.QSize(20, 20))
        self.pushButton_9.setObjectName("pushButton_9")
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_28.setFont(font)
        self.label_28.setAlignment(QtCore.Qt.AlignCenter)
        self.label_28.setObjectName("label_28")
        self.pushButton.raise_()
        self.pushButton_2.raise_()
        self.comboBox.raise_()
        self.dial.raise_()
        self.horizontalSlider.raise_()
        self.lineEdit.raise_()
        self.lineEdit_2.raise_()
        self.lineEdit_3.raise_()
        self.lineEdit_4.raise_()
        self.pushButton_3.raise_()
        self.pushButton_4.raise_()
        self.pushButton_5.raise_()
        self.pushButton_6.raise_()
        self.label_4.raise_()
        self.comboBox_2.raise_()
        self.pushButton_7.raise_()
        self.label_6.raise_()
        self.label_06.raise_()
        self.horizontalSlider_2.raise_()
        self.label_7.raise_()
        self.label_8.raise_()
        self.frame_2.raise_()
        self.frame.raise_()
        self.label_5.raise_()
        self.dial_2.raise_()
        self.label_9.raise_()
        self.lineEdit.setAlignment(Qt.AlignCenter)
        self.lineEdit_2.setAlignment(Qt.AlignCenter)
        self.lineEdit_3.setAlignment(Qt.AlignCenter)
        self.lineEdit_4.setAlignment(Qt.AlignCenter)
        self.lineEdit_5.setAlignment(Qt.AlignCenter)
        self.lineEdit_6.setAlignment(Qt.AlignCenter)
        self.lineEdit_7.setAlignment(Qt.AlignCenter)
        self.lineEdit_8.setAlignment(Qt.AlignCenter)
        self.lineEdit_9.setAlignment(Qt.AlignCenter)
        self.lineEdit_10.setAlignment(Qt.AlignCenter)
        self.lineEdit_11.setAlignment(Qt.AlignCenter)
        self.lineEdit_12.setAlignment(Qt.AlignCenter)
        self.lineEdit_13.setAlignment(Qt.AlignCenter)
        self.lineEdit_14.setAlignment(Qt.AlignCenter)
        self.lineEdit_15.setAlignment(Qt.AlignCenter)
        self.lineEdit_16.setAlignment(Qt.AlignCenter)
        self.lineEdit_17.setAlignment(Qt.AlignCenter)
        self.lineEdit_18.setAlignment(Qt.AlignCenter)
        self.lineEdit_21.setAlignment(Qt.AlignCenter)
        self.lineEdit_23.setAlignment(Qt.AlignCenter)
        self.lineEdit_24.setAlignment(Qt.AlignCenter)
        self.lineEdit.setText('0 m')
        self.lineEdit_2.setText('0 m')
        self.lineEdit_3.setText('0 m')
        self.lineEdit_4.setText('0 m')
        self.lineEdit_5.setText('500')
        self.lineEdit_6.setText('1000')
        self.lineEdit_7.setText('2000')
        self.lineEdit_8.setText('10')
        self.lineEdit_9.setText('18.2812')
        self.lineEdit_10.setText('NaN')
        self.lineEdit_11.setText('x=0 ;y=0 ;z=0')
        self.lineEdit_12.setText('36')
        self.lineEdit_13.setText('18.2812')
        self.lineEdit_14.setText('926')
        self.lineEdit_23.setText('43.7')
        self.lineEdit_24.setText('0.155')
        self.lineEdit_17.setText('0.8787')
        self.lineEdit_21.setText('181')
        self.lineEdit_18.setText('194')
        self.lineEdit_15.setText('0.04')
        self.lineEdit_16.setText('0')
        self.table.hide()
        self.table_2.hide()
        self.table_3.hide()
        self.table_4.hide()
        self.table_5.hide()
        self.table_6.hide()
        self.table_7.hide()
        self.table_8.hide()
        self.table_9.hide()
        self.table_10.hide()
        self.tables = []

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)
        self.comboBox.addItems(mainWindow.sensor_file_names)
        self.comboBox.currentIndexChanged.connect(lambda: self.comboActivity())
        self.dial.valueChanged.connect(lambda: self.dialer())
        self.tables.append(self.table)
        self.tables.append(self.table_2)
        self.tables.append(self.table_3)
        self.tables.append(self.table_4)
        self.tables.append(self.table_5)
        self.tables.append(self.table_6)
        self.tables.append(self.table_7)
        self.tables.append(self.table_8)
        self.tables.append(self.table_9)
        self.tables.append(self.table_10)

        self.comboBox.setCurrentIndex(0)
        self.comboBox_2.setCurrentIndex(0)
        for b in range(0,len(mainWindow.imported_sensor_names)):
            self.showTable(b,self.tables[b])
        self.comboBox_2.clear()
        self.comboBox_2.addItems(mainWindow.imported_sensor_names[0])

        self.table.show()

    def create_map_graph(self):
        generateMap.createSmallMap(self.horizontalLayout)
        def generate_datapoints(selection, x_t, y_t, z_t, azimuth, elevation, C_drag, Caliber_diameter, Caliber_mass,
                                air_density, muzzle_velocity):
            global X_pos, Y_pos, Z_pos, time, Vx_list, Vy_list, Vz_list, \
                X_pos_2, Y_pos_2, Z_pos_2, time_2, Vx_list_2, Vy_list_2, Vz_list_2, t_end
            g = 9.8066
            V0_z = muzzle_velocity * np.sin(elevation * 0.0174532925)
            V0_x = muzzle_velocity * np.cos(elevation * 0.0174532925) * np.cos(azimuth * 0.0174532925)
            V0_y = muzzle_velocity * np.cos(elevation * 0.0174532925) * np.sin(azimuth * 0.0174532925)
            Caliber_cross_section = np.pi * (Caliber_diameter * 0.001) ** 2
            V_terminal = (2 * Caliber_mass * g / (C_drag * air_density * Caliber_cross_section)) ** 0.5
            V_z = V0_z
            V_x = V0_x
            V_y = V0_y
            t_peak = 0
            while True:
                D_z = 0.5 * C_drag * air_density * Caliber_cross_section * V_z
                D_x = 0.5 * C_drag * air_density * Caliber_cross_section * V_x
                D_y = 0.5 * C_drag * air_density * Caliber_cross_section * V_y
                V_z = V_z - (g + (D_z / Caliber_mass))
                V_x = V_x - (D_x / Caliber_mass)
                V_y = V_y - (D_y / Caliber_mass)
                x_t = x_t + V_x
                y_t = y_t + V_y
                z_t = z_t + V_z
                if selection == 1:
                    X_pos.append(x_t)
                    Y_pos.append(y_t)
                    Z_pos.append(z_t)
                elif selection == 2:
                    X_pos_2.append(x_t)
                    Y_pos_2.append(y_t)
                    Z_pos_2.append(z_t)
                if V_z <= 0:
                    H_max = z_t
                    H_max_x = x_t
                    H_max_y = y_t
                    break
                else:
                    t_peak = t_peak + 1
            z_t = H_max
            x_t = H_max_x
            y_t = H_max_y
            t_end = t_peak
            while True:
                D_z = 0.5 * C_drag * air_density * Caliber_cross_section * V_z
                D_x = 0.5 * C_drag * air_density * Caliber_cross_section * V_x
                D_y = 0.5 * C_drag * air_density * Caliber_cross_section * V_y
                V_x = V_x - (D_x / Caliber_mass)
                V_y = V_y - (D_y / Caliber_mass)
                V_z = V_z + (g - (D_z / Caliber_mass))
                x_t = x_t + (V_x)
                y_t = y_t + (V_y)
                z_t = z_t - V_z
                if selection == 1:
                    X_pos.append(x_t)
                    Y_pos.append(y_t)
                    Z_pos.append(z_t)
                elif selection == 2:
                    X_pos_2.append(x_t)
                    Y_pos_2.append(y_t)
                    Z_pos_2.append(z_t)
                if V_z >= V_terminal:
                    V_z = V_terminal
                else:
                    pass
                if z_t <= 0:
                    break
                else:
                    t_end = t_end + 1

            if selection == 1:
                X_pos.reverse()
                Y_pos.reverse()
            else:
                X_pos_2.reverse()
                Y_pos_2.reverse()
        generate_datapoints(1, 0, 0, 0, 45, 25, 0.5, 155, 50, 1.225, 750)
        generate_datapoints(2, 0, 0, 0, 45, 35, 0.5, 155, 50, 1.225, 650)
        X_pos_n = np.asarray(X_pos)
        Y_pos_n = np.asarray(Y_pos)
        Z_pos_n = np.asarray(Z_pos)
        X_pos_2_n = np.asarray(X_pos_2)
        Y_pos_2_n = np.asarray(Y_pos_2)
        Z_pos_2_n = np.asarray(Z_pos_2)
        self.fig = plt.figure()
        self.canvas = FigureCanvas(self.fig)
        self.axes = self.fig.add_subplot(111, projection='3d')
        self.axes.xaxis._axinfo["grid"].update({"linewidth": 0.5, 'color': 'black', 'linestyle': 'dashed'})
        self.axes.yaxis._axinfo["grid"].update({"linewidth": 0.5, 'color': 'black'})
        self.axes.zaxis._axinfo["grid"].update({"linewidth": 0.5, 'color': 'black'})
        self.axes.w_xaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
        self.axes.w_yaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
        self.axes.w_zaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
        self.axes.plot(X_pos_n, Y_pos_n, Z_pos_n)
        colors = ['#7FBB87', '#D66272', '#BA86F8', '#F9B52E']
        count = 0
        self.axes.margins(x=0, y=-0.4)
        self.canvas.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding)
        self.canvas.updateGeometry()
        self.axes.mouse_init()
        self.axes.tick_params(axis='x', labelsize=5)
        self.axes.tick_params(axis='y', labelsize=5)
        self.axes.tick_params(axis='z', labelsize=5)
        self.axes.set_xlim(0, 50000)
        self.axes.set_ylim(0, 50000)
        self.axes.set_zlim(0, 20000)
        for i in [Z_pos[45], Z_pos[51], Z_pos[32], Z_pos[39]]:
            r = 10
            z0 = i  # To have the tangent at y=0
            y0 = Y_pos[Z_pos.index(i)]
            x0 = X_pos[Z_pos.index(i)]
            # Theta varies only between pi/2 and 3pi/2. to have a half-circle
            theta = np.linspace(0, 2 * np.pi, 201)
            a = 1.1
            color = colors[count]
            h = [45, 51, 32, 39]
            for j in range(11):
                z = (3 * r * np.cos(theta) + z0)  # y - y0 = r*cos(theta)
                y = 4 * r * np.sin(theta) + y0  # z - z0 = r*sin(theta)
                x = 4 * r * np.sin(-theta) + x0
                a = a - 0.1
                self.axes.plot(x, y, z, linestyle='--', alpha=a, c=color)
                r = r + 100  # To have the tangent at y=0
                y0 = Y_pos[Z_pos.index(Z_pos[h[count] - j])]
                x0 = X_pos[Z_pos.index(Z_pos[h[count] - j])]
            count = count + 1
        X_Sensor = [9000, 12000, 28000, 33000]
        Y_Sensor = [35000, 37500, 25000, 23000]
        Z_Sensor = [0, 0, 0, 0]
        txt = ['CASTLE-1', 'CASTLE-2', 'CASTLE-3', 'CASTLE-4']
        self.axes.scatter(xs=X_Sensor, ys=Y_Sensor, zs=Z_Sensor, s=10, c='blue')
        for k in range(4):
            self.axes.text(X_Sensor[k], Y_Sensor[k], Z_Sensor[k], txt[k], size=5, zorder=1, color='k')
        X_arrow = []
        Y_arrow = []
        Z_arrow = []
        for l in [Z_pos[51], Z_pos[45], Z_pos[32], Z_pos[39]]:
            X_arrow.append(X_pos[Z_pos.index(l)])
            Y_arrow.append(Y_pos[Z_pos.index(l)])
            Z_arrow.append(l)
        colors_2 = ['#D66272', '#7FBB87', '#BA86F8', '#F9B52E']
        for m in range(4):
            sketch_x = []
            sketch_y = []
            sketch_z = []
            sketch_x.append(X_arrow[m])
            sketch_y.append(Y_arrow[m])
            sketch_z.append(Z_arrow[m])
            sketch_x.append(X_Sensor[m])
            sketch_y.append(Y_Sensor[m])
            sketch_z.append(Z_Sensor[m])
        self.horizontalLayout_2.addWidget(self.canvas)

    def comboActivity(self):
        delays.clear()
        delays_total.clear()
        delay_ranking.clear()
        for k in self.tables:
            k.hide()
        self.comboBox_2.clear()
        self.comboBox_2.addItems(mainWindow.imported_sensor_names[self.comboBox.currentIndex()])
        self.tables[self.comboBox.currentIndex()].show()

    def second_smallest(self, numbers):
        m1 = m2 = float('inf')
        for x in numbers:
            if x <= m1:
                m1, m2 = x, m1
            elif x < m2:
                m2 = x
        return m2

    def showTable(self,i,table):
        self.create_map_graph()
        v_header = []
        delays = []
        table.setColumnCount(22)
        table.setRowCount(len(mainWindow.imported_events) * (len(mainWindow.imported_sensor_names[i]) + 1))
        table.setHorizontalHeaderLabels(
            ['Sensor Post Coordinates (lat,long)','Delay Between Neighbour Signals Impact (s)',
             'Delay Between Neighbour Signals Firing (s)',
             '3D SW - TOA (s)', 'Firing MB - TOA (s)', 'Impact - TOA (s)', '3D Shockwave Arrival Time',
             'Firing Signal Arrival Time', 'Impact Signal Arrival Time', 'Firing Coordinate (lat,long)',
             'Impact Coordinate (lat,long)', 'Bearing of Firing (°)', 'Bearing of Impact (°)', 'Name of Launch Point',
             'Name of Impact Point', 'Firing Time', 'Impact Time', 'Distance to Firing Point (meters)',
             'Distance to Impact Point (meters)', 'Distance Firing-Impact Point (meters)',
             'Average Velocity of Caliber (m/s)', 'Average Speed of Sound (m/s)'])
        differences = []
        sensor_times = []
        ordered_differences = []
        for a in range(0, len(mainWindow.imported_sensor_names[i])):
            sensor_times.append([])
        for m in range(0, len(mainWindow.imported_events)):
            for n in range(0, len(mainWindow.imported_sensor_names[i])):
                delays.append(
                    table_data_launch[0][i * len(mainWindow.imported_events) + m][
                        (5 + len(mainWindow.imported_sensor_names[i]) + n)])
                delays.append(
                    table_data_impact[0][i * len(mainWindow.imported_events) + m][
                        (5 + len(mainWindow.imported_sensor_names[i]) + n)])
                sensor_times[n].append(table_data_launch[0][i * len(mainWindow.imported_events) + m][
                        (5 + len(mainWindow.imported_sensor_names[i]) + n)])
                sensor_times[n].append(table_data_impact[0][i * len(mainWindow.imported_events) + m][
                        (5 + len(mainWindow.imported_sensor_names[i]) + n)])
                sensor_times[n].sort()
        delays_total.append(delays)
        header = table.horizontalHeader()
        total_differences.append(differences)
        total_ordered_diff.append(ordered_differences)
        for vheader in range(0,len(mainWindow.imported_events)):
            v_header.append(' ')
            for sensor in mainWindow.imported_sensor_names[i]:
                v_header.append(sensor)
        for head in range(0,len(header)):
            header.setSectionResizeMode(head, QtWidgets.QHeaderView.ResizeToContents)
        table.setVerticalHeaderLabels(v_header)

        table.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter |
                                                           Qt.Alignment(QtCore.Qt.TextWordWrap))
        table.verticalHeader().setDefaultAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter |
                                                         Qt.Alignment(QtCore.Qt.TextWordWrap))
        table.setStyleSheet("QHeaderView::section {background-color: #F8F4E2; height: 50 ; width: 100;}")
        self.comboBox_2.addItems(mainWindow.imported_sensor_names[i])
        try:
            sensor_info = []
            sensor_individual = []
            for m in range(0, len(mainWindow.imported_events)):
                for n in range(0, len(mainWindow.imported_sensor_names[i])):
                    sensor_individual.append(mainWindow.imported_sensor_coords[i][n])
                    sensor_individual.append('ordered_differences[(m+1)*2*n]')
                    sensor_individual.append('ordered_differences[(m+1)*2*n+1]')
                    sensor_individual.append(' # ')
                    sensor_individual.append(table_data_launch[0][i*len(mainWindow.imported_events)+m][(5 + len(mainWindow.imported_sensor_names[i]) + n)])
                    sensor_individual.append(table_data_impact[0][i*len(mainWindow.imported_events)+m][(5 + len(mainWindow.imported_sensor_names[i]) + n)])
                    sensor_individual.append(' # ')
                    sensor_individual.append(table_data_launch[0][m][5 + n])
                    sensor_individual.append(table_data_impact[0][m][5 + n])
                    sensor_individual.append(table_data_launch[0][m][1])
                    sensor_individual.append(table_data_impact[0][m][2])
                    sensor_individual.append(
                        bearing_sensor_launch[0][(m * len(mainWindow.imported_sensor_names[i])) + n])
                    sensor_individual.append(
                        bearing_sensor_impact[0][(m * len(mainWindow.imported_sensor_names[i])) + n])
                    sensor_individual.append(mainWindow.imported_events[m][0])
                    sensor_individual.append(mainWindow.imported_events[m][2])
                    sensor_individual.append(table_data_launch[0][m][3])
                    sensor_individual.append(table_data_impact[0][m][4])
                    sensor_individual.append(
                        dist_sensor_launch[0][(m * len(mainWindow.imported_sensor_names[i])) + n])
                    sensor_individual.append(
                        dist_sensor_impact[0][(m * len(mainWindow.imported_sensor_names[i])) + n])
                    sensor_individual.append(dist_event[0][m])
                    if mainWindow.imported_events[m][6] == 1:
                        sensor_individual.append(
                            (dist_event[0][m] / float(mainWindow.imported_events[m][7])))
                    else:
                        sensor_individual.append(mainWindow.imported_events[m][7])
                    sensor_individual.append(mainWindow.imported_events[m][8])
                    sensor_info.append(sensor_individual)
                    sensor_individual = []
                for k in range(0, len(mainWindow.imported_events)):
                    for j in range(0, len(mainWindow.imported_events[k])):
                        table.setSpan(j * (len(mainWindow.imported_sensor_names[i]) + 1), 0, 1, 22)
                        cell = QTableWidgetItem('                     Event ' + str(j + 1))
                        cell.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                        cell.setBackground(QtGui.QColor(255, 180, 110))
                        table.setItem(j * (len(mainWindow.imported_sensor_names[i]) + 1), 0, cell)

            for l in range(0, len(mainWindow.imported_events)):
                for a in range(0, len(mainWindow.imported_sensor_names[i])):
                    col = 0
                    for item in sensor_info[l * len(mainWindow.imported_sensor_names[i]) + a]:
                        cell_2 = QTableWidgetItem(str(item))
                        cell_2.setTextAlignment(Qt.AlignCenter)
                        table.setItem(l * (len(mainWindow.imported_sensor_names[i]) + 1) + a + 1, col, cell_2)
                        col = col + 1
        except IndexError:
            pass

    def releaseCheck(self):
        global distance, initial_point, degree, pos, last_distance, last_point
        state_left = win32api.GetKeyState(0x01)  # Left button down = 0 or 1. Button up = -127 or -128
        while True:
            a = win32api.GetKeyState(0x01)
            if a != state_left:  # Button state changed
                state_left = a
                if a < 0:
                    pass
                else:
                    for h in range(0,len(mainWindow.imported_events)):
                        for l in range(0, 21):
                            self.tables[self.comboBox.currentIndex()].item(((h*(self.comboBox_2.count()+1))+(self.comboBox_2.currentIndex() + 1)), l).setBackground(
                                QtGui.QColor(255, int(distance / 10), int(distance / 10)))
                    break
            time.sleep(0.001)

    def dialer(self):
        global distance, initial_point, degree, pos, last_distance, table_pointer
        value = self.dial.value()
        if value >= 0 and value < 180:
            degree = 180 - value
        elif value < 360 and value >= 180:
            degree = 360 - (value - 180)
        self.label_6.setText(str(distance)+'m')
        self.label_06.setText(str(degree)+'°')
        pos = str(pyautogui.position()).split('Point(x=')
        pos = pos[-1].split(',')
        y=pos[-1].replace('y=','')
        y= y.replace(')','')
        pos.pop(-1)
        pos.append(y)
        initial_point.append(pos)
        distance = 2*round(math.sqrt(abs(((int(initial_point[0][0]))-(int(pos[0])))**2+((int(initial_point[0][1]))-(int(pos[1])))**2)))
        if len(initial_point) > 1:
            initial_point = initial_point[:-1]
        else:
            pass
        if ((int(initial_point[0][1]))-(int(pos[1]))) >= 0:
            self.lineEdit.setText(str((int(initial_point[0][1]))-(int(pos[1]))) + ' m')
            self.lineEdit_4.setText('0 m')
        else:
            self.lineEdit.setText('0 m')
            self.lineEdit_4.setText(str(abs((int(initial_point[0][1])) - (int(pos[1])))) + ' m')
        if (int(initial_point[0][0])-(int(pos[0]))) >= 0:
            self.lineEdit_3.setText(str(int(initial_point[0][0])-(int(pos[0]))) + ' m')
            self.lineEdit_2.setText('0 m')
        else:
            self.lineEdit_2.setText(str(abs(int(initial_point[0][0]) - (int(pos[0])))) + ' m')
            self.lineEdit_3.setText('0 m')
        t1 = threading.Thread(target=self.releaseCheck)
        t1.start()

    def addTableRow(self,row_data,index):
        row = self.table.rowCount()
        self.table.setRowCount(row + 1)
        col = 0
        find_match = table_data_launch[0] + table_data_impact[0]
        delays = []
        for item in row_data:
            cell = QTableWidgetItem(str(item))
            cell.setTextAlignment(Qt.AlignCenter)
            if col > (4+len(mainWindow.imported_sensor_names[index])):
                for i in range(0,len(find_match)):
                    for j in range(0,len(find_match[i])):
                        if j == col and find_match[i][j] <= item:
                            delay = item - find_match[i][j]
                            delays.append(delay)
                            if self.second_smallest(delays) <= 2 and self.second_smallest(delays) > 0:
                                cell.setBackground(QtGui.QColor(255, 95, 95))
                            elif self.second_smallest(delays) > 2 and self.second_smallest(delays) <= 4:
                                cell.setBackground(QtGui.QColor(255, 145, 145))
                            elif self.second_smallest(delays) > 4 and self.second_smallest(delays) <= 10:
                                cell.setBackground(QtGui.QColor(255, 195, 195))
                            elif self.second_smallest(delays) > 10:
                                cell.setBackground(QtGui.QColor(219, 255, 221))
                            elif len(delays) == 1:
                                cell.setBackground(QtGui.QColor(219, 255, 221))
                            else:
                                pass
                        else:
                            pass
            else:
                pass
            delays.clear()
            self.table.setItem(row, col, cell)
            col = col + 1

    def addTableRow_2(self,row_data):
        row = self.table2.rowCount()
        self.table2.setRowCount(row + 1)
        col = 0
        delays = []
        for item in row_data:
            cell = QTableWidgetItem(str(item))
            cell.setTextAlignment(Qt.AlignCenter)
            for i in range(0, len(row_data)):
                if row_data[i] <= item:
                    delay = item - row_data[i]
                    delays.append(delay)
                    if self.second_smallest(delays) <= 2 and self.second_smallest(delays) > 0:
                        cell.setBackground(QtGui.QColor(255, 95, 95))
                    elif self.second_smallest(delays) > 2 and self.second_smallest(delays) <= 4:
                        cell.setBackground(QtGui.QColor(255, 145, 145))
                    elif self.second_smallest(delays) > 4 and self.second_smallest(delays) <= 10:
                        cell.setBackground(QtGui.QColor(255, 195, 195))
                    elif self.second_smallest(delays) > 10:
                        cell.setBackground(QtGui.QColor(219, 255, 221))
                    elif len(delays) == 1:
                        cell.setBackground(QtGui.QColor(219, 255, 221))
                    else:
                        pass
                else:
                    pass
            delays.clear()
            self.table2.setItem(row, col, cell)
            col = col + 1

    def exportExcel(self):
        global sensor_seconds, sensor_info
        file_filter = 'Data File (*.xlsx *.csv *.dat);; Excel File (*.xlsx *.xls)'
        response = QFileDialog.getSaveFileName(parent=None,
            caption='Select a data file',
            directory='mission.xlsx',
            filter=file_filter,
            initialFilter='Excel File (*.xlsx *.xls)'
        )
        try:
            if self.const == 1:
                columnHeaders = []
                for j in range(self.table.model().columnCount()):
                    columnHeaders.append(self.table.horizontalHeaderItem(j).text())
                df = pd.DataFrame(columns=columnHeaders)
                for row in range(self.table.rowCount()):
                    for col in range(self.table.columnCount()):
                        df.at[row, columnHeaders[col]] = self.table.item(row, col).text()
                df.to_excel(response[0])

            elif self.const == 2:
                columnHeaders_2 = []
                for i in range(self.table1.model().columnCount()):
                    columnHeaders_2.append(self.table1.horizontalHeaderItem(i).text())
                df_2 = pd.DataFrame(columns=columnHeaders_2)
                df_2.to_excel('mission.xlsx')
                wb = openpyxl.load_workbook('mission.xlsx')
                sheet = wb.active
                for j in range(0, len(calculateParameters.event_array)):
                    sheet.cell(row = (j * (len(calculateParameters.name_sensor) + 1) + 2), column = 5).value = 'Event ' + str(j + 1)
                for k in range(0, len(calculateParameters.event_array)):
                    for a in range(0, len(calculateParameters.name_sensor)):
                        col = 2
                        sheet.cell(row=(k * (len(calculateParameters.name_sensor) + 1) + a + 3), column=1).value = calculateParameters.name_sensor[a]
                        for item in sensor_info[k * len(calculateParameters.name_sensor) + a]:
                            sheet.cell(row=(k * (len(calculateParameters.name_sensor) + 1) + a + 3), column=col).value = str(item)
                            col = col + 1
                wb.save(response[0])

            else:
                wb = openpyxl.load_workbook('mission.xlsx')
                sheet = wb.active
                sheet.delete_rows(1, sheet.max_row )
                for i in range(0, len(calculateParameters.name_sensor)):
                    sheet.cell(row=i+1 , column=1).value = calculateParameters.name_sensor[i]
                    for j in range(0,len(sensor_seconds[i])):
                        sheet.cell(row=i+1, column=j+2).value = str(sensor_seconds[i][j])
                wb.save(response[0])
        except ValueError:
            pass


    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Event Analysis"))
        self.pushButton.setText(_translate("Dialog", "  Export Table"))
        self.pushButton_2.setText(_translate("Dialog", "  Save New\n"
                                                       " Deployment"))
        self.label_4.setText(_translate("Dialog", "Arrange Translation of Selected Sensor Post:"))
        self.pushButton_7.setText(_translate("Dialog", "   Back to First\n"
                                                       "   Deployment"))
        self.label_6.setText(_translate("Dialog", "0m"))
        self.label_06.setText(_translate("Dialog", "0°"))
        self.label_7.setText(_translate("Dialog", "Arrange Wind Direction &  Speed:"))
        self.label_8.setText(_translate("Dialog", "0 m/s"))
        self.label_9.setText(_translate("Dialog", "0°"))
        self.checkBox.setText(_translate("Dialog", "Disabled"))
        self.checkBox_2.setText(_translate("Dialog", "Diasbled"))
        self.label_11.setText(_translate("Dialog", "Sensor Resolution [m]:"))
        self.label_12.setText(_translate("Dialog", "Grid Resolution [m]:"))
        self.label_13.setText(_translate("Dialog", "Temperature [°C]:"))
        self.label_14.setText(_translate("Dialog", "Grid-Sensor Resolution [m]:"))
        self.label_15.setText(_translate("Dialog", "Azimuth [mils]:"))
        self.label_16.setText(_translate("Dialog", "Weapon Position [m]:"))
        self.label_17.setText(_translate("Dialog", "Weapon Ref. Point [UTM]:"))
        self.label_18.setText(_translate("Dialog", "Elevation [mils]:"))
        self.label_19.setText(_translate("Dialog", "Caliber Diameter [m]:"))
        self.label_20.setText(_translate("Dialog", "Attenuation Coefficient [dB/100m]:"))
        self.label_21.setText(_translate("Dialog", "Shockwave Source Level [dB SPL]:"))
        self.label_22.setText(_translate("Dialog", "Impact Position Height [m]:"))
        self.label_23.setText(_translate("Dialog", "Caliber Mass [kg]:"))
        self.label_24.setText(_translate("Dialog", "Muzzle Velocity [m/s]:"))
        self.label_25.setText(_translate("Dialog", "Spherical Wave Source Level [dB SPL]:"))
        self.label_26.setText(_translate("Dialog", "Caliber Lenght [m]:"))
        self.label_27.setText(_translate("Dialog", "Air Density [kg/m3]:"))
        self.label_28.setText(_translate("Dialog", "Advanced Settings"))
        self.pushButton_8.setText(_translate("Dialog", " Save Settings"))
        self.pushButton_9.setText(_translate("Dialog", "Default Settings"))


def createTable():
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    Dialog.exec_()

