from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QTableWidgetItem, QLabel, QFileDialog
import calculateParameters as Cp
import pandas as pd
import openpyxl
import mainWindow

table_data_launch = []
table_data_impact = []
dist_sensor_launch = []
dist_sensor_impact = []
dist_event = []
bearing_sensor_launch = []
bearing_sensor_impact = []
bearing_event = []
const = 1

class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(1161, 689)
        self.pushButton = QtWidgets.QPushButton(Dialog, clicked = lambda: self.exportExcel())
        self.pushButton.setGeometry(QtCore.QRect(70, 620, 91, 51))
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("save.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton.setIcon(icon)
        self.pushButton.setObjectName("pushButton")
        self.pushButton_2 = QtWidgets.QPushButton(Dialog, clicked = lambda: self.changeView())
        self.pushButton_2.setGeometry(QtCore.QRect(170, 620, 91, 51))
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("sort.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_2.setIcon(icon1)
        self.pushButton_2.setObjectName("pushButton_2")
        self.label = QLabel(Dialog)
        pixmap = QPixmap('colorbar.png')
        self.label.setPixmap(pixmap)
        self.label.setGeometry(QtCore.QRect(873, 620, 1, 1))
        self.label.resize(pixmap.width(), pixmap.height())
        self.label1 = QLabel(Dialog)
        self.label1.setGeometry(QtCore.QRect(870, 635, 1, 1))
        self.label1.resize(pixmap.width(), pixmap.height())
        self.label1.setObjectName("label1")
        self.table1 = QtWidgets.QTableWidget(Dialog)
        self.table1.setGeometry(QtCore.QRect(25, 21, 1111, 581))
        self.table1.setObjectName("tableWidget1")
        self.table2 = QtWidgets.QTableWidget(Dialog)
        self.table2.setGeometry(QtCore.QRect(25, 21, 1111, 581))
        self.table2.setObjectName("tableWidget2")
        self.table = QtWidgets.QTableWidget(Dialog)
        self.table.setGeometry(QtCore.QRect(25, 21, 1111, 581))
        self.table.setObjectName("tableWidget")
        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)
        h_header=['Sensor Post Coordinates (lat,long)', 'Firing Signal Arrival Time', 'Impact Signal Arrival Time',
                  'Firing Coordinate (lat,long)', 'Impact Coordinate (lat,long)', 'Bearing of Firing (°)',
                  'Bearing of Impact (°)', 'Name of Launch Point', 'Name of Impact Point', 'Firing Time',
                  'Impact Time', 'Distance to Firing Point (meters)', 'Distance to Impact Point (meters)',
                  'Distance Firing-Impact Point (meters)', 'Average Velocity of Caliber (m/s)',
                  'Average Speed of Sound (m/s)']
        v_header=['sensors']
        for header in mainWindow.imported_sensor_names:
            for name in header:
                v_header.append(str(name))
        self.table.setColumnCount(len(Cp.headers)+len(Cp.name_sensor)+len(Cp.name_sensor))
        self.table.setHorizontalHeaderLabels(Cp.headers+Cp.name_sensor+Cp.name_sensor)
        self.table.setStyleSheet("QHeaderView::section {background-color: #F8F4E2;}")
        self.table1.setColumnCount(len(h_header))
        self.table1.setRowCount(len(v_header))
        self.table1.setHorizontalHeaderLabels(h_header)
        self.table1.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter |
                                                           Qt.Alignment(QtCore.Qt.TextWordWrap))
        self.table1.verticalHeader().setDefaultAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter |
                                                         Qt.Alignment(QtCore.Qt.TextWordWrap))
        self.table1.setStyleSheet("QHeaderView::section {background-color: #F8F4E2; height: 50 ; width: 100;}")
        self.table1.setVerticalHeaderLabels(v_header)
        self.table2.setColumnCount(len(Cp.event_array)*2)
        self.table2.verticalHeader().setDefaultAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter |
                                                         Qt.Alignment(QtCore.Qt.TextWordWrap))
        self.table2.setStyleSheet("QHeaderView::section {background-color: #F8F4E2; height: 50; width: 100;}")
        self.createWholeInfo(0)
    def createWholeInfo(self,i):
        print(mainWindow.imported_events)
        print(mainWindow.imported_sensor_names)
        print(mainWindow.imported_sensor_coords)
        print(table_data_launch)
        print(table_data_impact)
        print(bearing_sensor_launch)
        print(bearing_sensor_impact)
        print(bearing_event)
        print(dist_sensor_launch)
        print(dist_sensor_impact)
        print(dist_event)
        try:
            single_seconds = []
            sensor_seconds = []
            sorter = []
            for x in range(0, len(mainWindow.imported_sensor_names[i])):
                for c in range(0, len(mainWindow.imported_events)):
                    single_seconds.append(
                        table_data_launch[0][c][(5 + len(mainWindow.imported_sensor_names[i]) + x)])
                    single_seconds.append(
                        table_data_impact[0][c][(5 + len(mainWindow.imported_sensor_names[i]) + x)])
                sorter = sorted(single_seconds)
                sensor_seconds.append(sorter)
                single_seconds = []
                sorter = []
            for y in range(0, len(mainWindow.imported_sensor_names[i])):
                self.addTableRow_2(sensor_seconds[y],i)
            self.table2.setVerticalHeaderLabels(mainWindow.imported_sensor_names[i])
            sensor_info = []
            sensor_individual = []
            for m in range(0, len(mainWindow.imported_events)):
                for n in range(0, len(mainWindow.imported_sensor_names[i])):
                    sensor_individual.append(mainWindow.imported_sensor_coords[i][n])
                    sensor_individual.append(table_data_launch[0][m][5 + n])
                    sensor_individual.append(table_data_impact[0][m][5 + n])
                    sensor_individual.append(table_data_launch[0][m][1])
                    sensor_individual.append(table_data_impact[0][m][2])
                    sensor_individual.append(
                        bearing_sensor_launch[0][(m * len(mainWindow.imported_sensor_names[i])) + n])
                    sensor_individual.append(
                        bearing_sensor_impact[0][(m * len(mainWindow.imported_sensor_names[i])) + n])
                    sensor_individual.append(mainWindow.imported_events[m][0])
                    sensor_individual.append(mainWindow.imported_events[m][2])
                    sensor_individual.append(table_data_launch[0][m][3])
                    sensor_individual.append(table_data_impact[0][m][4])
                    sensor_individual.append(
                        dist_sensor_launch[0][(m * len(mainWindow.imported_sensor_names[i])) + n])
                    sensor_individual.append(
                        dist_sensor_impact[0][(m * len(mainWindow.imported_sensor_names[i])) + n])
                    sensor_individual.append(dist_event[0][m])
                    if mainWindow.imported_events[m][6] == 1:
                        sensor_individual.append(
                            (dist_event[0][m] / float(mainWindow.imported_events[m][7])))
                    else:
                        sensor_individual.append(mainWindow.imported_events[m][7])
                    sensor_individual.append(mainWindow.imported_events[m][8])
                    sensor_info.append(sensor_individual)
                    sensor_individual = []
                for k in range(0, len(mainWindow.imported_events)):
                    for j in range(0, len(mainWindow.imported_events[k])):
                        self.table1.setSpan(j * (len(mainWindow.imported_sensor_names[i]) + 1), 0, 1, 16)
                        cell = QTableWidgetItem('Event ' + str(j + 1))
                        cell.setTextAlignment(Qt.AlignCenter)
                        cell.setBackground(QtGui.QColor(255, 180, 110))
                        self.table1.setItem(j * (len(mainWindow.imported_sensor_names[i]) + 1), 0, cell)

            for k in range(0, len(mainWindow.imported_events)):
                for a in range(0, len(mainWindow.imported_sensor_names[i])):
                    col = 0
                    for item in sensor_info[k * len(mainWindow.imported_sensor_names[i]) + a]:
                        cell_2 = QTableWidgetItem(str(item))
                        cell_2.setTextAlignment(Qt.AlignCenter)
                        self.table1.setItem(k * (len(mainWindow.imported_sensor_names[i]) + 1) + a + 1, col, cell_2)
                        col = col + 1
            for k in range(0, len(mainWindow.imported_events)):
                for l in table_data_launch[k]:
                    self.addTableRow(l,i)
                for b in table_data_impact[k]:
                    self.addTableRow(b,i)
        except IndexError:
            pass

    def changeView(self):
        global const
        if const == 0:
            self.table1.hide()
            self.table2.hide()
            self.table.show()
            const = 1
        elif const == 1:
            self.table.hide()
            self.table2.hide()
            self.table1.show()
            const = 2
        else:
            self.table.hide()
            self.table1.hide()
            self.table2.show()
            const = 0

    def second_smallest(self, numbers):
        m1 = m2 = float('inf')
        for x in numbers:
            if x <= m1:
                m1, m2 = x, m1
            elif x < m2:
                m2 = x
        return m2

    def addTableRow(self,row_data,index):
        row = self.table.rowCount()
        self.table.setRowCount(row + 1)
        col = 0
        find_match = table_data_launch[0] + table_data_impact[0]
        delays = []
        for item in row_data:
            cell = QTableWidgetItem(str(item))
            cell.setTextAlignment(Qt.AlignCenter)
            if col > (4+len(mainWindow.imported_sensor_names[index])):
                for i in range(0,len(find_match)):
                    for j in range(0,len(find_match[i])):
                        if j == col and find_match[i][j] <= item:
                            delay = item - find_match[i][j]
                            delays.append(delay)
                            if self.second_smallest(delays) <= 2 and self.second_smallest(delays) > 0:
                                cell.setBackground(QtGui.QColor(255, 95, 95))
                            elif self.second_smallest(delays) > 2 and self.second_smallest(delays) <= 4:
                                cell.setBackground(QtGui.QColor(255, 145, 145))
                            elif self.second_smallest(delays) > 4 and self.second_smallest(delays) <= 10:
                                cell.setBackground(QtGui.QColor(255, 195, 195))
                            elif self.second_smallest(delays) > 10:
                                cell.setBackground(QtGui.QColor(219, 255, 221))
                            elif len(delays) == 1:
                                cell.setBackground(QtGui.QColor(219, 255, 221))
                            else:
                                pass
                        else:
                            pass
            else:
                pass
            delays.clear()
            self.table.setItem(row, col, cell)
            col = col + 1

          
    def addTableRow_2(self,row_data,index):
        row = self.table2.rowCount()
        self.table2.setRowCount(row + 1)
        col = 0
        delays = []
        for item in row_data:
            cell = QTableWidgetItem(str(item))
            cell.setTextAlignment(Qt.AlignCenter)
            for i in range(0, len(row_data)):
                if row_data[i] <= item:
                    delay = item - row_data[i]
                    delays.append(delay)
                    if self.second_smallest(delays) <= 2 and self.second_smallest(delays) > 0:
                        cell.setBackground(QtGui.QColor(255, 95, 95))
                    elif self.second_smallest(delays) > 2 and self.second_smallest(delays) <= 4:
                        cell.setBackground(QtGui.QColor(255, 145, 145))
                    elif self.second_smallest(delays) > 4 and self.second_smallest(delays) <= 10:
                        cell.setBackground(QtGui.QColor(255, 195, 195))
                    elif self.second_smallest(delays) > 10:
                        cell.setBackground(QtGui.QColor(219, 255, 221))
                    elif len(delays) == 1:
                        cell.setBackground(QtGui.QColor(219, 255, 221))
                    else:
                        pass
                else:
                    pass
            delays.clear()
            self.table2.setItem(row, col, cell)
            col = col + 1

    def exportExcel(self):
        global sensor_seconds, sensor_info
        file_filter = 'Data File (*.xlsx *.csv *.dat);; Excel File (*.xlsx *.xls)'
        response = QFileDialog.getSaveFileName(parent=None,
            caption='Select a data file',
            directory='mission.xlsx',
            filter=file_filter,
            initialFilter='Excel File (*.xlsx *.xls)'
        )
        try:
            if self.const == 1:
                columnHeaders = []
                for j in range(self.table.model().columnCount()):
                    columnHeaders.append(self.table.horizontalHeaderItem(j).text())
                df = pd.DataFrame(columns=columnHeaders)
                for row in range(self.table.rowCount()):
                    for col in range(self.table.columnCount()):
                        df.at[row, columnHeaders[col]] = self.table.item(row, col).text()
                df.to_excel(response[0])

            elif self.const == 2:
                columnHeaders_2 = []
                for i in range(self.table1.model().columnCount()):
                    columnHeaders_2.append(self.table1.horizontalHeaderItem(i).text())
                df_2 = pd.DataFrame(columns=columnHeaders_2)
                df_2.to_excel('mission.xlsx')
                wb = openpyxl.load_workbook('mission.xlsx')
                sheet = wb.active
                for j in range(0, len(Cp.event_array)):
                    sheet.cell(row = (j * (len(Cp.name_sensor) + 1) + 2), column = 5).value = 'Event ' + str(j + 1)
                for k in range(0, len(Cp.event_array)):
                    for a in range(0, len(Cp.name_sensor)):
                        col = 2
                        sheet.cell(row=(k * (len(Cp.name_sensor) + 1) + a + 3), column=1).value = Cp.name_sensor[a]
                        for item in sensor_info[k * len(Cp.name_sensor) + a]:
                            sheet.cell(row=(k * (len(Cp.name_sensor) + 1) + a + 3), column=col).value = str(item)
                            col = col + 1
                wb.save(response[0])

            else:
                wb = openpyxl.load_workbook('mission.xlsx')
                sheet = wb.active
                sheet.delete_rows(1, sheet.max_row )
                for i in range(0, len(Cp.name_sensor)):
                    sheet.cell(row=i+1 , column=1).value = Cp.name_sensor[i]
                    for j in range(0,len(sensor_seconds[i])):
                        sheet.cell(row=i+1, column=j+2).value = str(sensor_seconds[i][j])
                wb.save(response[0])
        except ValueError:
            pass


    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Results of Event Analysis"))
        self.pushButton.setText(_translate("Dialog", "  Save"))
        self.pushButton_2.setText(_translate("Dialog", " Change \n   View"))
        self.label1.setText(_translate("Dialog", "  > 10s  |  > 4s  |  > 2s  |  > 0s"))

def createTable():
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    Dialog.exec_()

def initiateTable():
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.close()